import os
import gspread
import pandas as pd
from collections import Counter
from google.auth import default
from google.cloud import storage
from openpyxl.utils import get_column_letter
from gspread_dataframe import get_as_dataframe

# Load the sheet
sh = gspread.authorize(default()[0]).open_by_key("1NOaWXARQiSA6fquOtcouQPREPN4buYIf13tq_F6D9As")

# Load the bucket
BUCKET = "fc-secure-d99fbd65-eb27-4989-95b4-4cf559aa7d36"
bucket = storage.Client().bucket(BUCKET)
bucket.reload()

# Load the number of FASTQ files for each BCL/Index
fastqs = [blob.name for blob in bucket.list_blobs(prefix=f"fastqs") if blob.name.endswith(".fastq.gz")]
fastqs = Counter([(f.split("/")[1], f.split("/")[2].split("_S")[0]) for f in fastqs])
fastqs = pd.DataFrame([(k[0],k[1],v) for k,v in fastqs.items()], columns=["BCL", "Index", "FASTQs"])
assert not fastqs.duplicated(subset=["BCL", "Index"]).any()

def blob2link(blob):
    if not isinstance(blob, str):
        return pd.NA
    elif '\n' in blob:
        return blob
    else:
        return f'=HYPERLINK("https://storage.cloud.google.com/{BUCKET}/{blob}", "{os.path.basename(blob)}")'

################################################################################

# Load <Recon> worksheet
df = get_as_dataframe(sh.worksheet("Recon"))
ranges = {col: get_column_letter(df.columns.get_loc(col)+1)+"2" for col in ["QC", "summary", "FASTQs"]}
df = df.drop(columns=["QC", "summary", "FASTQs"])
dups = df.duplicated(subset=["BCL", "Index"])
assert not dups.any(), f"Recon sheet has duplicated BCL/Index pair:\n{df[dups]}"


# Load recon PDF blobs
recon = [blob.name for blob in bucket.list_blobs(prefix=f"recon") if blob.name.endswith(".pdf")]

QC = [(r.split("/")[1], r.split("/")[2], r) for r in recon if r.endswith("QC.pdf")]
summary = [(r.split("/")[1], r.split("/")[2], r) for r in recon if r.endswith("summary.pdf")]

QC = pd.DataFrame(QC, columns=["BCL", "Index", "QC"])
summary = pd.DataFrame(summary, columns=["BCL", "Index", "summary"])
summary = summary.groupby(['BCL', 'Index'], as_index=False).agg({'summary': lambda x: '\n'.join(x)})

assert not QC.duplicated(subset=["BCL", "Index"]).any()
assert not summary.duplicated(subset=["BCL", "Index"]).any()


# Update the sheet
QC = pd.merge(df, QC, on=["BCL", "Index"], how="left").fillna('')["QC"]
summary = pd.merge(df, summary, on=["BCL", "Index"], how="left").fillna('')["summary"]
recon_fastqs = pd.merge(df, fastqs, on=["BCL", "Index"], how="left").fillna('')["FASTQs"]

sh.worksheet("Recon").update(values=[[blob2link(v)] for v in QC], range_name=ranges["QC"], raw=False)
sh.worksheet("Recon").update(values=[[blob2link(v)] for v in summary], range_name=ranges["summary"], raw=False)
sh.worksheet("Recon").update(values=[[v] for v in recon_fastqs], range_name=ranges["FASTQs"], raw=False)

################################################################################

# Load <Slide-tags> worksheet
df = get_as_dataframe(sh.worksheet("Slide-tags"))
ranges = {col: get_column_letter(df.columns.get_loc(col)+1)+"2" for col in ["web_summary", "summary", "FASTQs"]}
df = df.drop(columns=["web_summary", "summary", "FASTQs"])

df.rename(columns={'RNAIndex': 'Index'}, inplace=True)
dups = df.duplicated(subset=["BCL", "Index"])
assert not dups.any(), f"Slide-tags sheet has duplicated BCL/RNAIndex pair:\n{df[dups]}"

# Load cellranger-count blobs
count = [blob.name for blob in bucket.list_blobs(prefix=f"gene-expression") if blob.name.endswith("/web_summary.html")]
count = [(c.split("/")[1], c.split("/")[2], blob2link(c)) for c in count]
count = pd.DataFrame(count, columns=["BCL", "Index", "web_summary"])
assert not count.duplicated(subset=["BCL", "Index"]).any()

# Load slide-tags blobs
tags = [blob.name for blob in bucket.list_blobs(prefix=f"slide-tags") if blob.name.endswith("/summary.pdf")]
tags = [(t.split("/")[1], t.split("/")[2], blob2link(t)) for t in tags]
tags = pd.DataFrame(tags, columns=["BCL", "Index", "summary"])
assert not tags.duplicated(subset=["BCL", "Index"]).any()

# Update the sheet
count = pd.merge(df, count, on=["BCL", "Index"], how="left").fillna('')["web_summary"]
tags = pd.merge(df, tags, on=["BCL", "Index"], how="left").fillna('')["summary"]
count_fastqs = pd.merge(df, fastqs, on=["BCL", "Index"], how="left").fillna('')["FASTQs"]

sh.worksheet("Slide-tags").update(values=[[v] for v in count], range_name=ranges["web_summary"], raw=False)
sh.worksheet("Slide-tags").update(values=[[v] for v in tags], range_name=ranges["summary"], raw=False)
sh.worksheet("Slide-tags").update(values=[[v] for v in count_fastqs], range_name=ranges["FASTQs"], raw=False)

'''
# Upload sheet
ws = sh.worksheet("Results")
set_with_dataframe(ws, df)

# Add horizontal lines
requests = []
for i in range(1, len(df)):
    if df.loc[i, 'BCL'] != df.loc[i-1, 'BCL']:
        requests.append({
            "updateBorders": {
                "range": {"sheetId": ws.id, "startRowIndex": i, "endRowIndex": i+1},
                "bottom": {"style": "SOLID"}
            }
        })

# Auto-resize all columns
requests.append({
    "autoResizeDimensions": {
        "dimensions": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": ws.col_count}
    }
})

# Submit changes
ws.spreadsheet.batch_update({"requests": requests})
'''